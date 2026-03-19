#!/usr/bin/env python3
"""
Sync target-companies.csv ↔ target-companies.xlsx.

Used by:
  - discovery_pipeline.py (after writing CSV)
  - apply_eval_results.py (after merging LLM evals)
  - web_prospecting.py (after merging prospecting results)

Direct usage:
  python3 scripts/core/target_companies_sync.py csv-to-xlsx
  python3 scripts/core/target_companies_sync.py xlsx-to-csv
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

DATA = Path(__file__).resolve().parents[2] / 'data'
CSV_PATH = DATA / 'target-companies.csv'
XLSX_PATH = DATA / 'target-companies.xlsx'

# Column widths by name — covers all known columns
COLUMN_WIDTHS = {
    'rank': 6,
    'company': 24,
    'website': 24,
    'careers_url': 42,
    'role_url': 28,
    'industry': 28,
    'size': 12,
    'stage': 16,
    'recent_funding': 30,
    'tech_signals': 34,
    'open_positions': 30,
    'last_checked': 14,
    'notes': 50,
    'role_family': 22,
    'source': 18,
    'location_detected': 20,
    'validation_status': 16,
    'exclusion_reason': 24,
    'llm_score': 12,
    'llm_rationale': 55,
    'llm_flags': 35,
    'llm_hard_pass': 14,
    'llm_hard_pass_reason': 30,
    'llm_evaluated_at': 22,
}

DEFAULT_COL_WIDTH = 18

# Row colors by validation_status
STATUS_FILLS = {
    'pass':       PatternFill('solid', fgColor='F0FDF4'),   # light green
    'watch_list': PatternFill('solid', fgColor='FFFBEB'),   # light amber
    'fail':       PatternFill('solid', fgColor='FEF2F2'),   # light red
}

# LLM score gradient: >=80 green, 60-79 yellow, <60 red
def llm_score_fill(score_str: str) -> PatternFill | None:
    try:
        s = float(score_str)
    except (TypeError, ValueError):
        return None
    if s >= 80:
        return PatternFill('solid', fgColor='DCFCE7')
    if s >= 60:
        return PatternFill('solid', fgColor='FEF9C3')
    return PatternFill('solid', fgColor='FEE2E2')


def csv_to_xlsx() -> None:
    if not CSV_PATH.exists():
        raise FileNotFoundError(f'missing csv: {CSV_PATH}')

    with CSV_PATH.open(encoding='utf-8') as f:
        rows = list(csv.reader(f))
    if not rows:
        raise RuntimeError('target-companies.csv is empty')

    headers = rows[0]
    data = rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Target Companies'

    # Header row
    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for c_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(1, c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Build column index lookups
    col_index = {name: idx + 1 for idx, name in enumerate(headers)}
    status_col = col_index.get('validation_status')
    llm_score_col = col_index.get('llm_score')

    # Data rows
    for r_idx, row in enumerate(data, 2):
        ws.append(row)

        # Determine row background by status
        status_val = ''
        if status_col and status_col <= len(row):
            status_val = row[status_col - 1]
        row_fill = STATUS_FILLS.get(status_val)

        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(r_idx, c_idx)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border
            if row_fill:
                cell.fill = row_fill

        # Override llm_score cell with score-based color
        if llm_score_col and llm_score_col <= len(row):
            score_val = row[llm_score_col - 1]
            score_fill = llm_score_fill(score_val)
            if score_fill:
                ws.cell(r_idx, llm_score_col).fill = score_fill
                ws.cell(r_idx, llm_score_col).font = Font(bold=True, size=10)

    # Column widths
    for c_idx, col_name in enumerate(headers, 1):
        width = COLUMN_WIDTHS.get(col_name, DEFAULT_COL_WIDTH)
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # Row height
    ws.row_dimensions[1].height = 30
    for r_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[r_idx].height = 60

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(XLSX_PATH)


def xlsx_to_csv() -> None:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f'missing xlsx: {XLSX_PATH}')

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Target Companies'] if 'Target Companies' in wb.sheetnames else wb.active

    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        vals = ['' if v is None else str(v) for v in row]
        if any(v != '' for v in vals):
            rows.append(vals)

    if not rows:
        raise RuntimeError('target-companies.xlsx has no rows')

    with CSV_PATH.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerows(rows)


if __name__ == '__main__':
    import sys
    mode = sys.argv[1] if len(sys.argv) > 1 else 'csv-to-xlsx'
    if mode == 'csv-to-xlsx':
        csv_to_xlsx()
        print(f'OK: wrote {XLSX_PATH}')
    elif mode == 'xlsx-to-csv':
        xlsx_to_csv()
        print(f'OK: wrote {CSV_PATH}')
    else:
        raise SystemExit('Usage: target_companies_sync.py [csv-to-xlsx|xlsx-to-csv]')
